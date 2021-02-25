import xlrd
import xlsxwriter
import openpyxl
import pandas as pd
import xlwt
import xlwings

source_xls = [r"\Users\M06429\Desktop\Portable Python-3.8.2 x64\test\test1.xlsx", r"\Users\M06429\Desktop\Portable Python-3.8.2 x64\test\test2.xlsx"]
target_xls = r"\Users\M06429\Desktop\Portable Python-3.8.2 x64\test\test5.xlsx"
data = []
for i in source_xls:
 wb = xlrd.open_workbook(i)
 for sheet in wb.sheets():
  for rownum in range(sheet.nrows):
   data.append(sheet.row_values(rownum))


workbook = xlsxwriter.Workbook(target_xls)
worksheet = workbook.add_worksheet()
font = workbook.add_format({"font_size":100})
for i in range(len(data)):
 for j in range(len(data[i])):
  worksheet.write(i, j, data[i][j], font)

workbook.close()

workbook = xlrd.open_workbook(r"\Users\M06429\Desktop\Portable Python-3.8.2 x64\test\test5.xlsx")

sheet = workbook.sheet_by_index(0)
data_list = []
row_list = []
nRows = sheet.nrows
nCols = sheet.ncols

for i in range(0, nRows):
    row_list = []
    for j in range(nCols):
        data_value = sheet.cell(i, j).value
        data_type = sheet.cell(i, j).ctype
        row_list.append(data_value)
    data_list.append(row_list)

#print(data_list)

from pandas import DataFrame
df = DataFrame(data_list)
df.to_excel(r"C:\Users\M06429\Desktop\123.xls", header = False , index = False)
print(df)









