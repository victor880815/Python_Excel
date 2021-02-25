import xlrd
import xlsxwriter
import openpyxl
import pandas as pd
import xlwt
import xlwings
from datetime import date, datetime

workbook = xlrd.open_workbook("Z:\MIS REPORT\案件回報\書翊\邱秀環2.xlsx")


sheet = workbook.sheet_by_index(0)
data_list = []
row_list = []
nRows = sheet.nrows
nCols = sheet.ncols
for i in range(0, nRows):
    row_list = []
    for j in range(nCols):
        # 獲取第i行，第j列的值
        data_value = sheet.cell(i, j).value
        # 獲取第i行，第j列的型別
        # ctype :  0 empty, 1 string ,2 number, 3 date, 4 boolean 5,error
        data_type = sheet.cell(i, j).ctype
        if data_type == 2:
            # 將字串轉為number
            data_value = str(int(data_value))
        if data_type == 3:
            # 對讀取資料表中日期列 進行格式化
            date_t = xlrd.xldate_as_tuple(data_value, workbook.datemode)
            data_value = date(*date_t[:3]).strftime('%Y-%m-%d')
        row_list.append(data_value)
    data_list.append(row_list)

#print(data_list)




fn = "Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx"

wb = openpyxl.load_workbook(fn)

wb.active = 0
ws = wb.active



for i in range(len(data_list)):
    ws.append((data_list[i][0],data_list[i][1],data_list[i][2],data_list[i][3],data_list[i][4],data_list[i][5],data_list[i][6],data_list[i][7],data_list[i][8],data_list[i][9]))



wb.save(fn)


df5 = pd.read_excel("Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx", sheet_name="H", index_col=None, header = 0,skiprows = 0)
#print(df5)

df6 = df5.drop_duplicates(subset=["申請書編號"], keep="last")
#print(df6)


df6.to_excel("Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx",sheet_name="H",index = False)

import openpyxl
wb3=openpyxl.load_workbook("Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx")
wb3.create_sheet(title='H回報',index=1)
wb3.save("Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx")

wb4=openpyxl.load_workbook("Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx")
wb4.create_sheet(title='CODE',index=2)
wb4.save("Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx")


workbook = xlrd.open_workbook("Z:\MIS REPORT\案件回報\書翊\CODE.xlsx")


sheet = workbook.sheet_by_index(0)
data = []
row_list = []
nRows = sheet.nrows
nCols = sheet.ncols
for i in range(0, nRows):
    row_list = []
    for j in range(nCols):
        # 獲取第i行，第j列的值
        data_value = sheet.cell(i, j).value
        # 獲取第i行，第j列的型別
        # ctype :  0 empty, 1 string ,2 number, 3 date, 4 boolean 5,error
        data_type = sheet.cell(i, j).ctype
        row_list.append(data_value)
    data.append(row_list)

print(data)

fn2 = "Z:\MIS REPORT\案件回報\未送二段原因回報_邱秀環.xlsx"
wb5 = openpyxl.load_workbook(fn2)

wb5.active = 2
ws5 = wb5.active

for i in range(len(data)):
    ws5.append((data[i][0],))

wb5.save(fn2)